from Digit_recognition import *
import tkinter as tk
from threading import Thread
import pandas as pd
import openpyxl as xl
import os
from PIL import Image, ImageDraw

# sets up the window
root = tk.Tk()
root.title("Digit recognition")
root.geometry("1400x410")
root.resizable(width=False, height=False)

params_outer_frame = tk.Frame(root)
params_outer_frame.grid(row=0, column=0)

tk.Label(params_outer_frame, text="Generate Parameters:", font=("Helvetica", 15)).grid(row=0, column=0)

params_frame = tk.Frame(params_outer_frame)
params_frame.grid(row=1, column=0)

accuracy_frame = tk.Frame(root)
accuracy_frame.grid(row=0, column=1)

accuracy_graph_frame = tk.Frame(accuracy_frame)
accuracy_graph_frame.grid(row=0, column=0)

accuracy_data_frame = tk.Frame(accuracy_frame)
accuracy_data_frame.grid(row=1, column=0)

image_frame = tk.Frame(root)
image_frame.grid(row=0, column=2)

image_frame_top = tk.Frame(image_frame)
image_frame_top.grid(row=0, column=0)

image_frame_mid = tk.Frame(image_frame)
image_frame_mid.grid(row=1, column=0)

image_frame_bot = tk.Frame(image_frame)
image_frame_bot.grid(row=2, column=0)

prediction_frame = tk.Frame(root)
prediction_frame.grid(row=0, column=3)

# creates a window displaying an error message
def error_msg(msg):
    error_win = tk.Toplevel()
    error_win.title(f"ERROR: {msg}")
    tk.Label(error_win, text=f"ERROR: {msg}").grid(row=0, column=0, padx=50, pady=50)

# displays the drawn image
def show_drawn_image(img_array):
    container = Figure(figsize=(2.5,2.5), dpi=100)
    graph = container.add_subplot(111)
    graph.imshow(img_array, cmap="gray")
    container.tight_layout()
    canvas = FigureCanvasTkAgg(container, master=image_frame_mid)
    canvas.draw()
    canvas.get_tk_widget().grid(row=0, column=0, padx=10, pady=10)

# allows user to draw an image which is converted into a format which can be forward propogated
def draw_and_convert_image():
    draw_win = tk.Toplevel()
    draw_win.title("Draw image")
    draw_win.resizable(width=False, height=False)

    canvas = tk.Canvas(draw_win, width=280, height=280, bg='black')
    canvas.grid(row=0, column=0)

    img = Image.new('L', (280, 280), 'black')
    draw = ImageDraw.Draw(img)

    prev_x = None
    prev_y = None

    def on_mouse_down(event):
        global prev_x, prev_y
        prev_x = event.x
        prev_y = event.y

    def on_mouse_drag(event):
        global prev_x, prev_y
        x, y = event.x, event.y
        brush_radius = 10
        canvas.create_oval(x - brush_radius, y - brush_radius, x + brush_radius, y + brush_radius, fill='white', outline='white')
        draw.ellipse([x - brush_radius, y - brush_radius, x + brush_radius, y + brush_radius], fill='white', outline='white')
        prev_x, prev_y = x, y

    def on_mouse_up(event):
        pass

    def convert_image():
        resized_img = img.resize((28, 28))
        img_array = np.array(resized_img).astype(np.uint8)

        draw_win.destroy()
        show_drawn_image(img_array)
        try:
            prediction, confidence = make_prediction_on_drawn_image(img_array.reshape(784, 1), w1, b1, w2, b2)
            prediction_label.config(text=f"Prediction: {prediction[0]}")
            actual_label.config(text="")
            conf_label.config(text=f"Confidence: {round(confidence[0]*100, 2)}% (2dp)")
        except NameError:
            error_msg("you have no loaded parameters")

    canvas.bind('<ButtonPress-1>', on_mouse_down)
    canvas.bind('<B1-Motion>', on_mouse_drag)
    canvas.bind('<ButtonRelease-1>', on_mouse_up)

    tk.Button(draw_win, text="Submit image", command=convert_image).grid(row=1, column=0)

# displays a pandas df of the parameters
def show_params(params, title):
    data_win = tk.Toplevel()
    data_win.title(title)
    tk.Label(data_win, text=pd.DataFrame(params)).pack()

def show_w1():
    try:
        show_params(w1, "Layer 1: Weights")#
    except NameError:
        error_msg("No loaded parameters")

def show_b1():
    try:
        show_params(b1, "Layer 1: Biases")
    except NameError:
        error_msg("No loaded parameters")

def show_w2():
    try:
        show_params(w2, "Layer 2: Weights")
    except NameError:
        error_msg("No loaded parameters")

def show_b2():
    try:
        show_params(b2, "Layer 2: Biases")
    except NameError:
        error_msg("No loaded parameters")

# returns a list of all files in a folder
def list_files_in_folder(folder_path):
    file_list = []
    for file_name in os.listdir(folder_path):
        if os.path.isfile(os.path.join(folder_path, file_name)):
            if file_name[-4:] == "xlsx":
                file_list.append(file_name[:len(file_name)-5])
    return file_list

# reads parameters from an excel file
def load_params(index):
    load_file_name = globals()[index].cget("text")
    wb = pd.ExcelFile(f"Digit recognition/parameters/{load_file_name}.xlsx")

    w1_df = wb.parse('w1', header=None)
    b1_df = wb.parse('b1', header=None)
    w2_df = wb.parse('w2', header=None)
    b2_df = wb.parse('b2', header=None)

    global w1, b1, w2, b2
    w1 = w1_df.values
    b1 = b1_df.values
    w2 = w2_df.values
    b2 = b2_df.values

    parameters_label.config(text=f"Parameters: {load_file_name}")
    load_win.destroy()

# saves parameters to an excel file
def save_params():
    create_wb = xl.Workbook()
    create_wb.remove(create_wb["Sheet"])
    sheet_names = ["w1", "b1", "w2", "b2"]
    for sheet in sheet_names:
        create_wb.create_sheet(sheet)
    create_wb.save(f"Digit recognition/parameters/{name.get()}.xlsx")

    w1_excel_df = pd.DataFrame(w1)
    b1_excel_df = pd.DataFrame(b1)
    w2_excel_df = pd.DataFrame(w2)
    b2_excel_df = pd.DataFrame(b2)

    workbook = xl.load_workbook(f"Digit recognition/parameters/{name.get()}.xlsx")
    sheet_names = ['w1', 'b1', 'w2', 'b2']
    dataframes = [w1_excel_df, b1_excel_df, w2_excel_df, b2_excel_df]
    for sheet_name, df in zip(sheet_names, dataframes):
        sheet = workbook[sheet_name]
        data = df.values.tolist()
        for row_index, row_data in enumerate(data, start=1):
            for col_index, cell_value in enumerate(row_data, start=1):
                sheet.cell(row=row_index, column=col_index, value=cell_value)
    workbook.save(f"Digit recognition/parameters/{name.get()}.xlsx")

    save_win.destroy()

# creates the window for loading parameters
def load_params_window():
    global load_win
    load_win = tk.Toplevel()
    load_win.title("Load parameters")

    file_list = list_files_in_folder("Digit recognition/parameters")

    if file_list == []:
        load_win.destroy()
        error_msg("No saved parameters")

    for index, file in enumerate(file_list):
        globals()[index] = tk.Button(load_win, text=file, command=lambda i=index: load_params(i))
        globals()[index].grid(row=index, column=0)

# creates the window for saving parameters
def save_params_window():
    global save_win
    save_win = tk.Toplevel()
    save_win.title("Save parameters")

    global name
    name = tk.Entry(save_win)
    name.grid(row=0, column=0)
    tk.Button(save_win, text="Save", command=save_params).grid(row=1, column=0)

# displays the output from the neural network
def show_prediction():
    prediction, actual, confidence = make_prediction(int(image_index_entry.get()), w1, b1, w2, b2)
    prediction_label.config(text=f"Prediction: {prediction[0]}")
    actual_label.config(text=f"Actual: {actual}")
    conf_label.config(text=f"Confidence: {round(confidence[0]*100, 2)}% (2dp)")

# shows image with the given index from the testing data set
def show_image():
    index = int(image_index_entry.get())
    if index >= 0:
        index = int(image_index_entry.get())
    else:
        index = 0
        image_index_entry.delete(0, 'end')
        image_index_entry.insert(0, "0")

    image = testing_images[:, index, None]
    image = image.reshape((28, 28)) * 255

    container = Figure(figsize=(2.5,2.5), dpi=100)
    graph = container.add_subplot(111)
    graph.imshow(image, cmap="gray")
    container.tight_layout()
    canvas = FigureCanvasTkAgg(container, master=image_frame_mid)
    canvas.draw()
    canvas.get_tk_widget().grid(row=0, column=0, padx=10, pady=10)
    
    try:
        show_prediction()
    except NameError:
        error_msg("No loaded parameters")

# allows the user to cycle left (-1 from the index) through the testing data
def left():
    current = int(image_index_entry.get())
    if current > 0:
        image_index_entry.delete(0, 'end')
        image_index_entry.insert(0, str(current-1))
    else:
        image_index_entry.delete(0, 'end')
        image_index_entry.insert(0, "0")
    show_image()

# allows the user to cycle right (+1 to the index) through the testing data  
def right():
    current = int(image_index_entry.get())
    if current >= 0:
        image_index_entry.delete(0, 'end')
        image_index_entry.insert(0, str(current+1))
    else:
        image_index_entry.delete(0, 'end')
        image_index_entry.insert(0, "0")
    show_image()

# obtains parameters by running the gradient dexcent with the neural network
def do_gradient_descent():
    global w1, b1, w2, b2
    try:
        w1, b1, w2, b2 = gradient_descent(images, labels, int(iterations_entry.get()), float(alpha_entry.get()), accuracy_data_frame, accuracy_graph_frame)
        parameters_label.config(text=f"Parameters: Generated")
    except ValueError:
        error_msg("Invalid inputs")

# allows the user to see live updates of accuracy
def gen_params():
    Thread(target=do_gradient_descent).start()

tk.Label(params_frame, text="Iterations:", width=20).grid(row=0, column=0, padx=10, pady=10)
tk.Label(params_frame, text="Learning rate:", width=20).grid(row=1, column=0, padx=10, pady=10)

iterations_entry = tk.Entry(params_frame, width=25)
iterations_entry.grid(row=0, column=1, padx=10, pady=10)
alpha_entry = tk.Entry(params_frame, width=25)
alpha_entry.grid(row=1, column=1, padx=10, pady=10)

tk.Button(params_frame, text="Generate Parameters", command=gen_params, width=45).grid(row=2, columnspan=2, padx=10, pady=10)

tk.Button(params_frame, text="Layer 1: Weights", command=show_w1, width=20).grid(row=3, column=0, padx=10, pady=10)
tk.Button(params_frame, text="Layer 1: Biases", command=show_b1, width=20).grid(row=3, column=1, padx=10, pady=10)
tk.Button(params_frame, text="Layer 2: Weights", command=show_w2, width=20).grid(row=4, column=0, padx=10, pady=10)
tk.Button(params_frame, text="Layer 2: Biases", command=show_b2, width=20).grid(row=4, column=1, padx=10, pady=10)

tk.Button(params_frame, text="Save parameters", width=45, command=save_params_window).grid(row=5, columnspan=2, padx=10, pady=10)
tk.Button(params_frame, text="Load parameters", width=45, command = load_params_window).grid(row=6, columnspan=2, padx=10, pady=10)

tk.Button(image_frame_top, text="Load image from test data", command=show_image).grid(row=0, column=0, padx=10, pady=10)
tk.Button(image_frame_top, text="Draw image", command=draw_and_convert_image).grid(row=0, column=1, padx=10, pady=10)

left_button = tk.Button(image_frame_bot, text="◀", command=left)
left_button.grid(row=0, column=0, padx=10, pady=10)
image_index_entry = tk.Entry(image_frame_bot)
image_index_entry.grid(row=0, column=1, padx=10, pady=10)
image_index_entry.insert(0, "0")
right_button = tk.Button(image_frame_bot, text="▶", command=right)
right_button.grid(row=0, column=2, padx=10, pady=10)

parameters_label = tk.Label(prediction_frame, text=f"Parameters:", font=("Helvetica", 20))
parameters_label.grid(row=0, column=0, padx=10, pady=10)
prediction_label = tk.Label(prediction_frame, text=f"Prediction:", font=("Helvetica", 20))
prediction_label.grid(row=1, column=0, padx=10, pady=10)
conf_label = tk.Label(prediction_frame, text=f"Confidence:", font=("Helvetica", 20))
conf_label.grid(row=2, column=0, padx=10, pady=10)
actual_label = tk.Label(prediction_frame, text=f"Actual:", font=("Helvetica", 20))
actual_label.grid(row=3, column=0, padx=10, pady=10)

root.mainloop()